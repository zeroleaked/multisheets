from errors import *

from bs4 import BeautifulSoup
import xlrd
import openpyxl
import csv
from io import StringIO
import pandas as pd
import re
import csv
import os

class Spreadsheet(object):

    def __init__(self, filename):
        self.filename = filename

        self.open_text()     

        self.parse_to_list()

        self.crop_table()

        self.equalize()

        headers = self.table[0]
        rows = self.table[1:]
        
        self.df = pd.DataFrame(rows, columns=headers)

    def parse_to_list(self):
        if self.file_type == "HTML":
            self.html_parse()

        elif self.file_type == "XML":
            self.xml_parse()

        elif self.file_type == "XLS" or self.file_type == "XLSX":
            self.xl_parse()

        else:
            self.csv_parse()

    
    def open_text(self):
        try:
            f = open(self.filename, "rb")
            wb = xlrd.open_workbook(file_contents=f.read())
            sheet = wb.sheet_by_index(0)
            self.table = []
            for row in range(sheet.nrows):
                self.table.append([])
                for cell in sheet.row_values(row):
                    if cell != '':
                        self.table[-1].append(cell)
            if not self.table:
                raise ExcelDecodeError
            self.file_type = "XLS"
            f.close()

        except xlrd.biffh.XLRDError as e:
            if "xlsx" in str(e):
                path, source_filename = os.path.split(self.filename)
                new_filename = os.path.splitext(self.filename)[0] + ".xlsx"
                os.rename(self.filename, new_filename)
                self.filename = new_filename

                wb = openpyxl.load_workbook(filename=self.filename)
                ws = wb[wb.sheetnames[0]]
                rows_iter = ws.iter_rows(max_col = ws.max_column, max_row = ws.max_row)
                table = [[cell.value for cell in row] for row in rows_iter]
                for row in table:
                    def isEmpty(x):
                        if x is None:
                            return True
                        elif isinstance(x, str):
                            if not x.strip():
                                return True
                        else:
                            return False
                    while row and isEmpty(row[-1]):
                        row.pop()
                self.table = table
                self.file_type = "XLSX"

            elif "BOF" in str(e):
                file = open(self.filename, "r")
                self.raw = file.read()
                file.close()

                self.soup = BeautifulSoup(self.raw, 'xml')
                self.tr0s = self.soup.find_all('Row')
                if self.tr0s:
                    self.file_type = "XML"
                    return
                    
                self.soup = BeautifulSoup(self.raw, 'html.parser')
                self.tr0s = self.soup.find_all('tr')   
                if self.tr0s:
                    self.file_type = "HTML"
                    return

                # find delimiter
                self.file_type = "CSV"
                nl = self.raw.index("\n")
                nl = self.raw.index("\n", nl)
                check = self.raw.index("|", 0, nl)
                if check != -1:
                    self.delimiter = "|"
                    return
                check = self.raw.index(",", 0, nl)
                if check != -1:
                    self.delimiter = ","
                    return
                check = self.raw.index(";", 0, nl)
                if check != -1:
                    self.delimiter = ";"
                    return
                
                raise FileTypeError # file cannot be read

    def html_parse(self):
        self.table = []
        for tr0 in self.tr0s:
            tr1s = tr0.find_all('tr')
            if len(tr1s) == 0:
                tds = tr0.find_all('td')
                row = []
                for td in tds:
                    row.append(td.get_text())
                if row:
                    self.table.append(row)

        if not self.table:
            raise EmptyTableError

    def xml_parse(self):
        self.table = []
        for tr0 in self.tr0s:
            # tr1s = tr0.find_all('Row')
            # if len(tr1s) == 0:
            tds = tr0.find_all('Cell')
            row = []
            for td in tds:
                row.append(td.get_text())
            if row:
                self.table.append(row)
        
        if not self.table:
            raise EmptyTableError

    def xl_parse(self):
        for i in range(len(self.table)):
            for j in range(len(self.table[i])):
                if isinstance(self.table[i][j], float):
                    self.table[i][j] = int(self.table[i][j])
        
        if not self.table:
            raise EmptyTableError

    def csv_parse(self):
        f = StringIO(self.raw)
        reader = csv.reader(f, delimiter=self.delimiter)
        self.table = []
        for row in reader:
            self.table.append(row)
        
        if not self.table:
            raise EmptyTableError
    
    def start_stop_index(self):
        n_cell = 0
        for row in self.table:
            n_cell += len(row)
        threshold = n_cell/len(self.table)

        start = -1
        for i in range(len(self.table)):
            if len(self.table[i]) >= threshold:
                start = i
                break
        if start == -1:
            raise StartNotFound

        stop = -1
        for i in range(start, len(self.table)):
            if len(self.table[i]) <= threshold:
                stop = i - 1
                break
        if stop == -1:
            stop = len(self.table) - 1

        return start, stop

    def crop_table(self):
        start, stop = self.start_stop_index()

        self.table = self.table[start: stop+1]

    def equalize(self):
        for i in range(len(self.table)):
            error = len(self.table[i]) - len(self.table[0])
            if error < 0:
                for j in range(-error):
                    self.table[i].append("")
            if error > 0:
                raise TableCropError

    def define_indexes(self, number_index, subtotal_index, details_indexes):
        self.number_index = number_index
        self.subtotal_index = subtotal_index
        self.details_indexes = details_indexes 

    def calculate_total(self):
        def str_to_int(x):
            if isinstance(x, int):
                return x
            out = int("".join(re.findall(r'\d+', x)))
            return out
        def int_to_str(x):
            out = ""
            bare = str(x)
            for i in range( len(bare)-1 , -1, -1 ):
                out = bare[i] + out
                if (len(bare) - i) % 3 == 0 and i > 0:
                    out = "." + out
            return out

        subtotal_int = self.df.iloc[:,self.subtotal_index].map(str_to_int)
        ppn_int = (subtotal_int * 0.1).astype(int)
        total_int = subtotal_int + ppn_int
        self.df["__PPN__"] = ppn_int.map(int_to_str)
        self.df["__Total__"] = total_int.map(int_to_str)
        self.df["__Total_Int__"] = total_int
        

    def threshold(self, cutoff):
        self.df = self.df[self.df["__Total__"] > cutoff]
        print(len(self.df), "nomor telepon memiliki tagihan lebih dari", cutoff)

    def match_mails(self, contacts_xls_path):
        contacts_df = pd.read_excel(contacts_xls_path, engine='openpyxl', dtype="string").dropna(axis=0, how="all").dropna(axis=1, how="all")
        number_label = self.df.columns[self.number_index]
        self.df = self.df.merge(contacts_df, left_on=number_label, right_on="No telepon", how="left", validate="many_to_many", suffixes=("_x", None))


        unmatched = self.df[pd.isnull(self.df["Email korporat"])]
        
        if len(unmatched):
            # with pd.option_context("display.max_rows", None, "display.max_columns", None)
            print(len(unmatched), "nomor telepon tidak memiliki alamat email")
            print("Data kontak :", contacts_xls_path)
            
            answer = input("Lewatkan email yang tidak ada? (Y/N) : ")
            if answer.lower() != "y":
                print("Batal")
                exit()
            else:
                self.df = self.df[self.df["Email korporat"].notna()]

    def extract_dict(self):
        extracts = []
        for index, row in self.df.iterrows():
            new_data = {
                "Nomor Telepon" : row["No telepon"],
                "Unit" : row["Unit"],
            }
            for detail in self.details_indexes:
                new_data[detail["label"]] = row.iloc[detail["index"]]
            new_data["Subtotal"] = row.iloc[self.subtotal_index]
            new_data["PPN"] = row["__PPN__"]
            new_data["Total"] = row["__Total__"]

            found = False
            for i in range( len(extracts) ):
                if extracts[i]["email"] == row["Email korporat"]:
                    found = True
                    extracts[i]["data"].append(new_data)
            
            if not found:
                new_email = {
                    "email" : row["Email korporat"],
                    "jabatan" : row["Senior Manager"],
                    "data" : [new_data]
                }

                extracts.append(new_email)
        return extracts

    def export_csv(self, date_str):

        path, source_filename = os.path.split(self.filename)
        csv_filename = os.path.join(path, "[DB_FORMATTED]" + os.path.splitext(source_filename)[0] + ".csv")

        number_label = self.df.columns[self.number_index]
        minimal_df = self.df[[ number_label, "__Total_Int__" ]].copy()
        minimal_df["empty"] = "NULL"
        minimal_df["date"] = date_str
        minimal_df = minimal_df[["empty", number_label, "date", "__Total_Int__", "empty" ]]
        minimal_df.columns = ["id", "notel", "date", "total", "created"]
        minimal_df.to_csv(csv_filename, index=False, header=False)
        print("CSV disimpan di :", csv_filename)
        self.minimal_df = minimal_df

if __name__ == "__main__":

    # path20 = "../raw_sample/2020.xls"
    # path21b = "../raw_sample/2021b.xls"
    # path21a = "../raw_sample/2021a.xls"
    # patho = "out.csv"
    path = "/home/zir/Dev/telephone/multisheets/sample.csv"
    path_contacts = "/home/zir/Dev/telephone/multisheets/sample.xlsx"
    spreadsheet = Spreadsheet(path)
    spreadsheet.define_indexes(2, [])
    spreadsheet.match_emails(path_contacts)
    print(spreadsheet.df)